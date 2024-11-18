from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_interval
from openpyxl.styles import (
    Alignment, Font
)

import os
import logging
from datetime import datetime

from functions import send_message_box, SMBOX_ICON_TYPE

class CExcelLog:
    _full_patch = None
    _file_name = None
    _file_name_in_logging = None
    _on_day_folder_name = None
    _BASE_FOLDER_NAME = "sn_logging_folder"

    # СТИЛЬ ШРИФТА
    _FONT_HEADER = Font(
        name='Calibri',
        size=30,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    _FONT_DATA = Font(
        name='Arial',
        size=20,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    _alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0
    )
    @classmethod
    def get_full_patch(cls) -> str: return cls._full_patch

    @classmethod
    def get_scanned_count(cls, vender: str, model: str) -> int:
        fpatch = cls.get_full_patch()
        if fpatch is None or not len(fpatch):
            cls.update_folder_name_vars()
            fpatch = cls.get_full_patch()

        if os.path.isfile(fpatch) is True:
            wb = load_workbook(fpatch)
            ws = wb.active

            # Укажите индексы столбцов для сравнения (например, A и B)
            col_a_index = 3  # Это первый столбец (A)
            col_b_index = 4  # Это второй столбец (B)

            count_finded = 0
            # Чтение значений из столбцов
            for row in range(1, ws.max_row + 1):  # Начинаем с первой строки
                in_cell_vender = ws.cell(row=row, column=col_a_index).value
                in_cell_model = ws.cell(row=row, column=col_b_index).value
                if in_cell_vender == vender and in_cell_model == model:
                    count_finded += 1

            return count_finded
        return 0


    @classmethod
    def update_folder_name_vars(cls):
        current_time = datetime.now()
        cls._on_day_folder_name = f"{current_time.day:02}_{current_time.month:02}_{current_time.year}"
        cls._file_name = f"scanned_log_{current_time.day:02}_{current_time.month:02}_{current_time.year}.xlsx"
        cls._full_patch = f"{cls._BASE_FOLDER_NAME}/{cls._on_day_folder_name}/{cls._file_name}"
        cls._file_name_in_logging = (f"{cls._BASE_FOLDER_NAME}/{cls._on_day_folder_name}/scanned_log_"
                                     f"{current_time.day:02}_{current_time.month:02}_{current_time.year}.log")

    @classmethod
    def print_log(cls, vender: str, model: str, sn: str) -> bool:
        current_time = datetime.now()

        if cls._on_day_folder_name is None:
            cls.update_folder_name_vars()
            # print(cls._on_day_folder_name)
            # print(cls._file_name)
            # print(cls._full_patch)

        full_time = f'{current_time.hour:02}:{current_time.minute:02}:{current_time.second:02}'
        full_date = f'{current_time.day:02}.{current_time.month:02}.{current_time.year}'

        try:
            if not os.path.isdir(cls._BASE_FOLDER_NAME):
                os.mkdir(cls._BASE_FOLDER_NAME + "/")

            if not os.path.isdir(f"{cls._BASE_FOLDER_NAME}/{cls._on_day_folder_name}"):
                os.mkdir(f"{cls._BASE_FOLDER_NAME}/{cls._on_day_folder_name}" + "/")

            logging.basicConfig(level=logging.INFO, filename=cls._file_name_in_logging, filemode="a",
                                format="%(asctime)s %(levelname)s %(message)s")

            if os.path.exists(cls._full_patch) is False:
                wb = Workbook()
                ws = wb.active
                ws.title = "Сканировка устройств"
                ws.append(("Дата:", "Время:", "Vendor:", "Model:", "SN:"))

                # задаём ширину и фонт для шапки
                cell_range = ws['A1':'E1']
                for i in cell_range:
                    for i2 in i:
                        letter_adress = i2.coordinate

                        ws[letter_adress].font = cls._FONT_HEADER
                        ws[letter_adress].alignment = cls._alignment

                interval = get_column_interval("A", "E")
                for item in interval:
                    ws.column_dimensions[item].width = 60
            else:
                wb = load_workbook(cls._full_patch)
                ws = wb.active

            ws.append((full_date, full_time,
                       vender, model, sn))

            for cell in ws:  # Проходим по ячейкам последней добавленной строки
                for item in cell:
                    letter_adress = item.coordinate
                    ws[letter_adress].font = cls._FONT_DATA  # Применяем шрифт к каждой ячейке
                    ws[letter_adress].alignment = cls._alignment

            wb.save(cls._full_patch)
            wb.close()

        except Exception as err:
            logging.warning(f"SN SuccessFull Scanned: Vendor: {vender} | Model: {model} | SN: {sn}")
            send_message_box(icon_style=SMBOX_ICON_TYPE.ICON_ERROR,
                             text=f"Не могу записать данные!\n\n"
                                  f"Ошибка: '{err}'",
                             title="Критическая ошибка",
                             variant_yes="Ок", variant_no="", callback=None)

            return False

        logging.info(f"SN SuccessFull Scanned: Vendor: {vender} | Model: {model} | SN: {sn}")

        return True
